"""
Listing Batch Optimizer - 批量标题描述优化模块
支持两种模式：
1. 数据库商品批量优化
2. Shopee 批量更新 Excel 模板读写
"""

import re
import sys
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast
from zipfile import ZIP_DEFLATED, ZipFile

from xml.etree import ElementTree as ET

import psycopg2
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WORKSPACE = Path('/root/.openclaw/workspace-e-commerce')
SCRIPTS_DIR = WORKSPACE / 'scripts'
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

sys.path.insert(0, '/root/.openclaw/workspace-e-commerce/config')
from llm_caller import call_llm_with_fallback_meta  # type: ignore[reportMissingImports]
from multisite_config import load_market_bundle, normalize_site_context  # type: ignore[reportMissingImports]


ProductData = Dict[str, Any]

DB_CONFIG = {
    'host': 'localhost',
    'database': 'ecommerce_data',
    'user': 'superuser',
    'password': 'Admin123!',
}

DEFAULT_FORBIDDEN_TERMS = [
    '现货', '現貨', '热销', '熱銷', '限量', '免运', '免運', '包邮', '包郵',
    '库存', '庫存', '当天寄出', '當天寄出', '24小時出貨', '24小时出货',
    '立即购买', '立即購買', '现在下单', '現在下單', '抢购', '搶購', '卖点', '賣點',
]

DEFAULT_SITE_LANGUAGE_MAP = {
    'shopee_tw': 'zh-Hant',
    'shopee_ph': 'en',
    'shopee_id': 'id',
    'shopee_th': 'th',
    'shopee_vn': 'vi',
    'shopee_my': 'en/ms',
    'shopee_sg': 'en',
    'shopee_br': 'pt',
    'shopee_mx': 'es',
}

EXCEL_COLUMN_KEYS = {
    'global_id': ('全球商品ID', 'ps_tmpl_mt_update_title_product_id'),
    'product_name': ('全球商品名称', 'ps_tmpl_mt_update_title_product_name'),
    'product_description': ('Global SKU description', 'ps_tmpl_mt_update_title_product_description'),
    'parent_sku': ('全球商品货号', 'ps_tmpl_mt_update_title_parent_sku'),
    'failure_reason': ('失败原因', 'et_title_reason'),
}

PANE_VALUE_FIXES = {
    'bottom_left': 'bottomLeft',
    'bottom_right': 'bottomRight',
    'top_left': 'topLeft',
    'top_right': 'topRight',
}


def normalize_text(value: Any) -> str:
    text = '' if value is None else str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def strip_markdown_fences(text: str) -> str:
    stripped = normalize_text(text)
    if stripped.startswith('```'):
        stripped = re.sub(r'^```[a-zA-Z0-9_-]*\n?', '', stripped)
        stripped = re.sub(r'\n?```$', '', stripped)
    return stripped.strip()


def remove_forbidden_terms(text: str, terms: Optional[List[str]] = None) -> str:
    cleaned = text
    for term in terms or DEFAULT_FORBIDDEN_TERMS:
        cleaned = cleaned.replace(term, '')
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
    return cleaned.strip()


def extract_keywords(text: str, limit: int = 8) -> List[str]:
    normalized = normalize_text(text).lower()
    raw_tokens = re.findall(r'[\w\u4e00-\u9fff]{2,}', normalized)
    stop_words = {
        '收納', '收纳', '商品', '全球', 'basic', 'info', 'description', 'product',
        '現貨', '现货', '熱銷', '热销', 'sku',
    }

    keywords: List[str] = []
    for token in raw_tokens:
        if token in stop_words or token.isdigit():
            continue
        if token not in keywords:
            keywords.append(token)
        if len(keywords) >= limit:
            break
    return keywords


def missing_keywords_in_head(description: str, title: str, limit: int = 4, head_length: int = 220) -> List[str]:
    # Keep this helper deterministic so validation and repair use the same rule.
    keywords = extract_keywords(title, limit=limit)
    desc_head = normalize_text(description)[:head_length].lower()
    return [word for word in keywords if word.lower() not in desc_head]


class ConsistencyChecker:
    """标题描述一致性检查器"""

    def __init__(self):
        self.conn: Any = psycopg2.connect(
            host=DB_CONFIG['host'],
            database=DB_CONFIG['database'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
        )

    def close(self) -> None:
        self.conn.close()

    def check_one(self, product: ProductData) -> Tuple[bool, List[str]]:
        title = normalize_text(product.get('optimized_title') or product.get('title'))
        desc = normalize_text(product.get('optimized_description'))
        issues: List[str] = []

        if not title or not desc:
            return False, ['标题或描述为空']

        if len(desc) < 100:
            return False, ['描述过短']

        title_words = extract_keywords(title, limit=5)
        desc_start = desc[:300].lower()
        missing = [word for word in title_words if word.lower() not in desc_start]
        if missing:
            issues.append(f'标题关键词未在描述前300字出现: {missing}')

        if '✨' in title and '✨' not in desc[:120]:
            issues.append('标题有✨但描述开头无✨')
        if '【' in title and '【' not in desc[:120]:
            issues.append('标题有【】但描述开头无【】')

        for word in title_words[:2]:
            if word.lower() not in desc[:150].lower():
                issues.append(f'品类词"{word}"未在描述前150字出现')

        return len(issues) == 0, issues

    def check_all_products(self, min_desc_len: int = 100) -> List[Dict[str, Any]]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT product_id_new, alibaba_product_id, title,
                   optimized_title, optimized_description
            FROM products
            WHERE (optimized_description IS NOT NULL AND optimized_description != '')
              AND LENGTH(optimized_description) >= %s
            ORDER BY updated_at DESC
            """,
            (min_desc_len,),
        )

        issues_list: List[Dict[str, Any]] = []
        for row in cur.fetchall():
            product: ProductData = {
                'item_no': row[0],
                'alibaba_id': row[1],
                'title': row[2],
                'optimized_title': row[3],
                'optimized_description': row[4],
            }
            is_consistent, issues = self.check_one(product)
            if not is_consistent:
                issues_list.append({'product': product, 'issues': issues})

        cur.close()
        return issues_list


class ListingBatchOptimizer:
    """批量 Listing 优化器，支持数据库模式和 Excel 模板模式。"""

    def __init__(self):
        self.conn: Any = psycopg2.connect(
            host=DB_CONFIG['host'],
            database=DB_CONFIG['database'],
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
        )
        self.last_llm_call: Dict[str, Any] = {}
        self.desc_prompt_template = self._read_text_file(
            '/root/.openclaw/workspace-e-commerce/config/prompts/desc_prompt_v3.md'
        )
        self.title_prompt_template = self._read_text_file(
            '/root/.openclaw/workspace-e-commerce/config/prompts/title_prompt_v3.md'
        )
        self.sku_prompt_template = self._read_text_file(
            '/root/.openclaw/workspace-e-commerce/config/prompts/sku_name_prompt_v1.md'
        )
        self.default_forbidden_terms = list(DEFAULT_FORBIDDEN_TERMS)
        self.last_prompt_trace: Dict[str, Any] = {}

    @staticmethod
    def _read_text_file(path: str) -> Optional[str]:
        try:
            return Path(path).read_text(encoding='utf-8')
        except Exception as exc:
            print(f'加载文件失败 {path}: {exc}')
            return None

    def close(self) -> None:
        self.conn.close()

    def _default_listing_language(self, site_code: str | None) -> str:
        normalized = normalize_text(site_code).lower()
        return DEFAULT_SITE_LANGUAGE_MAP.get(normalized, 'zh-Hant')

    @staticmethod
    def _as_dict(value: Any) -> Dict[str, Any]:
        return cast(Dict[str, Any], value) if isinstance(value, dict) else {}

    def _profile_prompt_source(self, profile: Dict[str, Any]) -> str:
        metadata = self._as_dict(profile.get('metadata'))
        if metadata.get('is_multisite_baseline'):
            return 'baseline_profile'
        return 'prompt_profile'

    def _set_last_prompt_trace(
        self,
        *,
        mode: str,
        prompt: str,
        prompt_source: str,
        template_version: str | None,
        rendered_variables: Dict[str, Any],
        bundle: Dict[str, Any],
    ) -> None:
        site_context = self._as_dict(bundle.get('site_context'))
        content_policy = self._as_dict(bundle.get('content_policy'))
        prompt_profile = self._as_dict(bundle.get('prompt_profile'))
        self.last_prompt_trace = {
            'mode': mode,
            'prompt': prompt,
            'prompt_source': prompt_source,
            'template_version': template_version or 'legacy',
            'rendered_variables': rendered_variables,
            'resolved': {
                'site_code': site_context.get('site_code'),
                'listing_language': self._listing_language(bundle),
                'content_policy_code': content_policy.get('content_policy_code'),
                'prompt_profile_code': prompt_profile.get('prompt_profile_code'),
            },
        }

    def get_last_prompt_trace(self) -> Dict[str, Any]:
        return dict(self.last_prompt_trace)

    def _resolve_prompt_template(self, mode: str, bundle: Dict[str, Any]) -> Tuple[Optional[str], str, str | None]:
        content_policy = self._as_dict(bundle.get('content_policy'))
        prompt_profile = self._as_dict(bundle.get('prompt_profile'))
        metadata = self._as_dict(content_policy.get('metadata'))
        prompt_metadata = self._as_dict(prompt_profile.get('metadata'))
        base_template = normalize_text(content_policy.get('prompt_base_template'))
        title_variant = normalize_text(content_policy.get('prompt_title_variant')) or normalize_text(metadata.get('prompt_title_variant'))
        desc_variant = normalize_text(content_policy.get('prompt_desc_variant')) or normalize_text(metadata.get('prompt_desc_variant'))
        profile_source = self._profile_prompt_source(prompt_profile)
        template_version = normalize_text(prompt_metadata.get('template_version')) or normalize_text(metadata.get('template_version')) or 'legacy'

        if mode == 'title':
            profile_template = normalize_text(prompt_profile.get('title_template')) or normalize_text(prompt_metadata.get('title_template'))
            if profile_template:
                return profile_template, profile_source, template_version
            if base_template or title_variant:
                legacy_template = f'{base_template}\n\n{title_variant}'.strip() if base_template and title_variant else (title_variant or base_template)
                return legacy_template or None, 'content_policy_legacy', template_version
            if self.title_prompt_template:
                return self.title_prompt_template, 'file_fallback', 'multisite-v1.0'
            return None, 'hardcoded_fallback', 'legacy'

        if mode == 'description':
            profile_template = normalize_text(prompt_profile.get('description_template')) or normalize_text(prompt_metadata.get('description_template'))
            if profile_template:
                return profile_template, profile_source, template_version
            if base_template or desc_variant:
                legacy_template = f'{base_template}\n\n{desc_variant}'.strip() if base_template and desc_variant else (desc_variant or base_template)
                return legacy_template or None, 'content_policy_legacy', template_version
            if self.desc_prompt_template:
                return self.desc_prompt_template, 'file_fallback', 'multisite-v1.0'
            return None, 'hardcoded_fallback', 'legacy'

        profile_template = normalize_text(prompt_profile.get('sku_name_template')) or normalize_text(prompt_metadata.get('sku_name_template'))
        if profile_template:
            return profile_template, profile_source, template_version
        if self.sku_prompt_template:
            return self.sku_prompt_template, 'file_fallback', 'multisite-v1.0'
        return None, 'hardcoded_fallback', 'legacy'

    def _resolve_runtime_bundle(self, product: Optional[ProductData] = None) -> Dict[str, Any]:
        product = product or {}
        site_context = cast(Dict[str, Any], normalize_site_context(product))
        try:
            return cast(
                Dict[str, Any],
                load_market_bundle(self.conn, market_code=site_context.get('market_code'), site_code=site_context.get('site_code')),
            )
        except Exception:
            return {
                'site_context': site_context,
                'market_config': {},
                'shipping_profile': {},
                'fee_profile': {},
                'content_policy': {},
                'prompt_profile': {},
            }

    def _active_forbidden_terms(self, bundle: Dict[str, Any]) -> List[str]:
        content_policy = self._as_dict(bundle.get('content_policy'))
        configured_terms = cast(List[Any], content_policy.get('forbidden_terms_json') or [])
        merged_terms = list(self.default_forbidden_terms)
        for term in configured_terms:
            text = normalize_text(term)
            if text and text not in merged_terms:
                merged_terms.append(text)
        return merged_terms

    def _render_prompt_template(self, template: str, variables: Dict[str, Any]) -> str:
        rendered = template
        for key, value in variables.items():
            rendered = rendered.replace(f'{{{key}}}', normalize_text(value))
        return rendered

    def _listing_language(self, bundle: Dict[str, Any]) -> str:
        site_context = self._as_dict(bundle.get('site_context'))
        content_policy = self._as_dict(bundle.get('content_policy'))
        site_code = site_context.get('site_code')
        return str(content_policy.get('listing_language') or site_context.get('listing_language') or self._default_listing_language(site_code)).strip()

    @contextmanager
    def _open_workbook(self, source: Path):
        sanitized_copy = self._sanitize_xlsx_for_openpyxl(source)
        try:
            workbook = load_workbook(sanitized_copy)
            yield workbook
        finally:
            if sanitized_copy != source and sanitized_copy.exists():
                sanitized_copy.unlink()

    def _sanitize_xlsx_for_openpyxl(self, source: Path) -> Path:
        namespace = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        needs_fix = False

        with ZipFile(source) as reader:
            for name in reader.namelist():
                if not name.startswith('xl/worksheets/') or not name.endswith('.xml'):
                    continue
                root = ET.fromstring(reader.read(name))
                for pane in root.findall('.//main:pane', namespace):
                    active_pane = pane.attrib.get('activePane')
                    if active_pane in PANE_VALUE_FIXES:
                        needs_fix = True
                        break
                if needs_fix:
                    break

        if not needs_fix:
            return source

        temp_file = tempfile.NamedTemporaryFile(suffix=source.suffix, delete=False)
        temp_path = Path(temp_file.name)
        temp_file.close()

        with ZipFile(source) as reader, ZipFile(temp_path, 'w', ZIP_DEFLATED) as writer:
            for item in reader.infolist():
                data = reader.read(item.filename)
                if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                    root = ET.fromstring(data)
                    modified = False
                    for pane in root.findall('.//main:pane', namespace):
                        active_pane = pane.attrib.get('activePane')
                        fixed_value = PANE_VALUE_FIXES.get(active_pane) if active_pane else None
                        if fixed_value:
                            pane.set('activePane', fixed_value)
                            modified = True
                    if modified:
                        data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                writer.writestr(item, data)

        return temp_path

    def _fetch_product_by_item_no(self, item_no: str) -> Optional[ProductData]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT product_id_new, alibaba_product_id, title, skus,
                   optimized_title, optimized_description
            FROM products
            WHERE product_id_new = %s
            """,
            (item_no,),
        )
        row = cur.fetchone()
        cur.close()
        if not row:
            return None
        return {
            'item_no': row[0],
            'alibaba_id': row[1],
            'title': row[2],
            'skus': row[3],
            'optimized_title': row[4],
            'optimized_description': row[5],
        }

    def _fetch_product_for_excel_row(self, item_no: str, global_id: str) -> Optional[ProductData]:
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT product_id_new, alibaba_product_id, title, skus,
                   optimized_title, optimized_description
            FROM products
            WHERE (%s != '' AND product_id_new = %s)
               OR (%s != '' AND alibaba_product_id = %s)
            ORDER BY updated_at DESC
            LIMIT 1
            """,
            (item_no, item_no, global_id, global_id),
        )
        row = cur.fetchone()
        cur.close()
        if not row:
            return None
        return {
            'item_no': row[0],
            'alibaba_id': row[1],
            'title': row[2],
            'skus': row[3],
            'optimized_title': row[4],
            'optimized_description': row[5],
        }

    def _update_product_content(self, item_no: str, content: ProductData) -> None:
        cur = self.conn.cursor()
        cur.execute(
            """
            UPDATE products
            SET optimized_title = %s,
                optimized_description = %s,
                optimization_version = COALESCE(optimization_version, 0) + 1,
                updated_at = NOW()
            WHERE product_id_new = %s
            """,
            (content['optimized_title'], content['optimized_description'], item_no),
        )
        self.conn.commit()
        cur.close()

    def _extract_product_info(self, product: ProductData) -> Dict[str, Any]:
        title = normalize_text(product.get('title') or product.get('optimized_title'))
        raw_skus = product.get('skus')
        skus: List[Any] = cast(List[Any], raw_skus) if isinstance(raw_skus, list) else []

        title_words = extract_keywords(title, limit=8)
        hot_words = '、'.join(title_words[:6]) or title
        features = '、'.join(title_words[:5]) or title

        sku_labels: List[str] = []
        for sku in skus[:5]:
            if isinstance(sku, dict):
                sku_dict = cast(Dict[str, Any], sku)
                label = normalize_text(sku_dict.get('sku_name') or sku_dict.get('name'))
                if label:
                    sku_labels.append(label)

        sku_summary = ' / '.join(sku_labels[:3])
        material_tokens = [
            word for word in title_words
            if any(key in word for key in ['鋼', '钢', '木', '布', '塑', '膠', '胶', '鐵', '铁', 'pet'])
        ]

        scenarios = '客廳、臥室、廚房、浴室'
        if any(word in title for word in ['廚房', '厨房', '水槽']):
            scenarios = '廚房、水槽下方、櫥櫃內部'
        elif any(word in title for word in ['浴室', '鏡櫃', '镜柜']):
            scenarios = '浴室、鏡櫃、洗手台周邊'

        return {
            'product_name': title,
            'original_title': title,
            'hot_search_words': hot_words,
            'features': features,
            'scenarios': scenarios,
            'material': '、'.join(material_tokens[:3]) or '依商品材質特性撰寫',
            'sku_count': len(skus),
            'attributes': sku_summary or hot_words,
        }

    def _build_title_prompt(self, product: ProductData, info: Dict[str, Any], bundle: Dict[str, Any]) -> str:
        title = normalize_text(product.get('title') or product.get('optimized_title'))
        title_keywords = '、'.join(extract_keywords(title, limit=6)) or title
        listing_language = self._listing_language(bundle)
        title_template, prompt_source, template_version = self._resolve_prompt_template('title', bundle)
        site_context = self._as_dict(bundle.get('site_context'))
        variables: Dict[str, Any] = {
            'original_title': info['original_title'],
            'attributes': info['attributes'],
            'hot_search_words': title_keywords,
            'listing_language': listing_language,
            'site_code': site_context.get('site_code') or '',
        }
        if title_template:
            prompt = self._render_prompt_template(title_template, variables)
            self._set_last_prompt_trace(mode='title', prompt=prompt, prompt_source=prompt_source, template_version=template_version, rendered_variables=variables, bundle=bundle)
            return prompt

        if listing_language.lower().startswith('en'):
            prompt = (
                f'Original product title: {info["original_title"]}\n'
                f'Product attributes: {info["attributes"]}\n'
                f'Search keywords: {title_keywords}\n\n'
                'Generate a concise Shopee Philippines English title between 40 and 80 characters. Avoid banned marketing terms.'
            )
            self._set_last_prompt_trace(mode='title', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables=variables, bundle=bundle)
            return prompt

        prompt = (
            f'商品原始标题：{info["original_title"]}\n'
            f'商品属性：{info["attributes"]}\n'
            f'热搜关键词：{title_keywords}\n\n'
            '请输出一个台湾繁体中文 Shopee 标题，长度 40-55 字，避免禁用词。'
        )
        self._set_last_prompt_trace(mode='title', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables=variables, bundle=bundle)
        return prompt

    def _build_desc_prompt(self, product: ProductData, info: Dict[str, Any], bundle: Dict[str, Any]) -> str:
        title = normalize_text(product.get('optimized_title') or product.get('title'))
        title_keywords = '、'.join(extract_keywords(title, limit=6)) or title
        listing_language = self._listing_language(bundle)
        desc_template, prompt_source, template_version = self._resolve_prompt_template('description', bundle)
        site_context = self._as_dict(bundle.get('site_context'))
        variables: Dict[str, Any] = {
            'product_name': info['product_name'],
            'material': info['material'],
            'features': info['features'],
            'scenarios': info['scenarios'],
            'hot_search_words': info['hot_search_words'],
            'listing_language': listing_language,
            'site_code': site_context.get('site_code') or '',
        }
        if desc_template:
            prompt = self._render_prompt_template(desc_template, variables)
            prompt += f'\n\n## 关键词一致性要求\n描述前200字必须自然包含这些关键词：{title_keywords}'
            self._set_last_prompt_trace(mode='description', prompt=prompt, prompt_source=prompt_source, template_version=template_version, rendered_variables={**variables, 'title_keywords': title_keywords}, bundle=bundle)
            return prompt

        if listing_language.lower().startswith('en'):
            prompt = (
                f'Product name: {info["product_name"]}\n'
                f'Material: {info["material"]}\n'
                f'Features: {info["features"]}\n'
                f'Usage scenarios: {info["scenarios"]}\n'
                f'Keywords: {title_keywords}\n\n'
                'Write an English Shopee Philippines product description with clear sections. The first 200 characters must naturally include the title keywords and avoid banned marketing terms.'
            )
            self._set_last_prompt_trace(mode='description', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables={**variables, 'title_keywords': title_keywords}, bundle=bundle)
            return prompt

        prompt = (
            f'商品名称：{info["product_name"]}\n'
            f'商品材质：{info["material"]}\n'
            f'商品特点：{info["features"]}\n'
            f'使用场景：{info["scenarios"]}\n'
            f'关键词：{title_keywords}\n\n'
            '请输出台湾繁体中文商品描述，七段式结构，前200字必须包含关键词，避免禁用词。'
        )
        self._set_last_prompt_trace(mode='description', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables={**variables, 'title_keywords': title_keywords}, bundle=bundle)
        return prompt

    def _call_llm_text(self, prompt: str, max_tokens: int, timeout: int) -> Optional[str]:
        messages = [{'role': 'user', 'content': prompt}]
        print('[LLM] 使用 Doubao + DeepSeek fallback')
        result = cast(Dict[str, Any], call_llm_with_fallback_meta(messages, max_tokens=max_tokens, timeout=timeout))
        self.last_llm_call = dict(result)
        return cast(Optional[str], result.get('text'))

    def get_last_llm_call(self) -> Dict[str, Any]:
        return dict(self.last_llm_call)

    def _sanitize_title(self, title: str, forbidden_terms: Optional[List[str]] = None) -> str:
        cleaned = strip_markdown_fences(title)
        cleaned = remove_forbidden_terms(cleaned, forbidden_terms)
        cleaned = cleaned.replace('\n', ' ')
        cleaned = re.sub(r'\s{2,}', ' ', cleaned)
        return cleaned.strip(' -_|，。,；;')

    def _sanitize_description(self, description: str, forbidden_terms: Optional[List[str]] = None) -> str:
        cleaned = strip_markdown_fences(description)
        cleaned = remove_forbidden_terms(cleaned, forbidden_terms)
        return normalize_text(cleaned)

    def _reinforce_description_head(self, description: str, title: str, forbidden_terms: Optional[List[str]] = None, listing_language: str = 'zh-Hant') -> str:
        cleaned = self._sanitize_description(description, forbidden_terms)
        missing = missing_keywords_in_head(cleaned, title)
        if not missing:
            return cleaned

        # Put the repair line at the very beginning so long first paragraphs
        # cannot push required title keywords beyond the validation window.
        if listing_language.lower().startswith('en'):
            focus_line = f'Key highlights: {", ".join(missing)} for quick product understanding.'
        else:
            focus_line = f'【重點整理】{"、".join(missing)}，一次掌握商品用途與收納特色。'
        return f'{focus_line}\n{cleaned}'.strip()

    def _validate_title(self, title: str, forbidden_terms: Optional[List[str]] = None, bundle: Optional[Dict[str, Any]] = None) -> Tuple[bool, str]:
        if not title:
            return False, '标题为空'
        bundle = bundle or {}
        content_policy = self._as_dict(bundle.get('content_policy'))
        rules = self._as_dict(content_policy.get('validation_rule_set'))
        min_length = int(rules.get('title_min_length') or 10)
        max_length = int(rules.get('title_max_length') or 60)
        if len(title) < min_length:
            return False, '标题过短'
        if len(title) > max_length:
            return False, '标题过长'
        for term in forbidden_terms or self.default_forbidden_terms:
            if term in title:
                return False, f'标题含禁用词: {term}'
        return True, ''

    def _validate_description(self, description: str, title: str, forbidden_terms: Optional[List[str]] = None, bundle: Optional[Dict[str, Any]] = None) -> Tuple[bool, str]:
        if not description:
            return False, '描述为空'
        bundle = bundle or {}
        content_policy = self._as_dict(bundle.get('content_policy'))
        rules = self._as_dict(content_policy.get('validation_rule_set'))
        min_length = int(rules.get('description_min_length') or 300)
        if len(description) < min_length:
            return False, '描述过短'
        for term in forbidden_terms or self.default_forbidden_terms:
            if term in description:
                return False, f'描述含禁用词: {term}'
        missing = missing_keywords_in_head(description, title)
        if missing:
            return False, f'描述前200字缺少标题关键词: {missing}'
        return True, ''

    def _sanitize_sku_listing_name(self, sku_name: str, max_length: int = 30, forbidden_terms: Optional[List[str]] = None) -> str:
        cleaned = strip_markdown_fences(sku_name)
        cleaned = remove_forbidden_terms(cleaned, forbidden_terms)
        cleaned = cleaned.replace('（', '(').replace('）', ')')
        cleaned = re.sub(r'\s+', ' ', cleaned).strip(' -_|，。,；;')
        if len(cleaned) > max_length:
            cleaned = cleaned[:max_length].rstrip(' -_|，。,；;')
        return cleaned

    def _derive_dimension_hint(self, sku: Dict[str, Any]) -> str:
        dims = [sku.get('package_length'), sku.get('package_width'), sku.get('package_height')]
        if all(isinstance(value, (int, float)) and float(value) > 0 for value in dims):
            numeric_dims = [float(cast(float, value)) for value in dims if isinstance(value, (int, float))]
            length, width, height = cast(Tuple[float, float, float], tuple(numeric_dims))
            return f"{int(length)}x{int(width)}x{int(height)}cm"

        source_text = normalize_text(
            sku.get('size')
            or sku.get('sku_name')
            or sku.get('name')
            or ''
        )
        match = re.search(r'(\d+(?:\.\d+)?\s*[xX×*＊]\s*\d+(?:\.\d+)?\s*[xX×*＊]\s*\d+(?:\.\d+)?\s*(?:cm|厘米|公分)?)', source_text)
        if match:
            return match.group(1).replace('×', 'x').replace('*', 'x').replace('＊', 'x')
        return ''

    def _fallback_sku_listing_name(self, product: ProductData, sku: Dict[str, Any], max_length: int = 30) -> str:
        raw_name = normalize_text(sku.get('shopee_sku_name') or sku.get('sku_name') or sku.get('name') or sku.get('size'))
        dimension_hint = self._derive_dimension_hint(sku)

        if not raw_name:
            raw_name = '標準規格'

        candidate = raw_name
        if dimension_hint and dimension_hint not in candidate and len(candidate) + len(dimension_hint) + 1 <= max_length:
            candidate = f'{candidate} {dimension_hint}'

        candidate = self._sanitize_sku_listing_name(candidate, max_length=max_length)
        return candidate or '標準規格'

    def _validate_sku_listing_name(self, sku_name: str, existing_names: List[str], max_length: int = 30, forbidden_terms: Optional[List[str]] = None) -> Tuple[bool, str]:
        if not sku_name:
            return False, 'SKU 名称为空'
        if len(sku_name) > max_length:
            return False, 'SKU 名称超长'
        if sku_name in existing_names:
            return False, 'SKU 名称重复'
        if sku_name.isdigit():
            return False, 'SKU 名称不能是纯数字'
        for term in forbidden_terms or self.default_forbidden_terms:
            if term in sku_name:
                return False, f'SKU 名称含禁用词: {term}'
        return True, ''

    def _ensure_unique_sku_name(self, candidate: str, existing_names: List[str], max_length: int = 30) -> str:
        if candidate not in existing_names:
            return candidate
        base = candidate[: max_length - 2].rstrip()
        suffix = 2
        while True:
            deduped = f'{base}{suffix}'
            deduped = deduped[:max_length]
            if deduped not in existing_names:
                return deduped
            suffix += 1

    def _generate_sku_listing_name(self, product: ProductData, sku: Dict[str, Any], existing_names: List[str], max_length: int = 30, bundle: Optional[Dict[str, Any]] = None) -> str:
        bundle = bundle or self._resolve_runtime_bundle(product)
        forbidden_terms = self._active_forbidden_terms(bundle)
        listing_language = self._listing_language(bundle)
        fallback_name = self._fallback_sku_listing_name(product, sku, max_length=max_length)
        title = normalize_text(product.get('optimized_title') or product.get('title'))
        description = normalize_text(product.get('optimized_description') or product.get('description'))
        raw_name = normalize_text(sku.get('sku_name') or sku.get('name') or sku.get('size'))
        dimension_hint = self._derive_dimension_hint(sku)
        custom_sku_template, prompt_source, template_version = self._resolve_prompt_template('sku', bundle)
        site_context = self._as_dict(bundle.get('site_context'))
        variables: Dict[str, Any] = {
            'product_title': title,
            'product_description': description[:600],
            'original_sku_name': raw_name,
            'dimension_hint': dimension_hint or 'none',
            'listing_language': listing_language,
            'site_code': site_context.get('site_code') or '',
            'max_length': str(max_length),
        }

        if custom_sku_template:
            prompt = self._render_prompt_template(custom_sku_template, variables)
            self._set_last_prompt_trace(mode='sku', prompt=prompt, prompt_source=prompt_source, template_version=template_version, rendered_variables=variables, bundle=bundle)
        elif listing_language.lower().startswith('en'):
            prompt = (
                'You are a Shopee Philippines SKU naming assistant.\n'
                'Goal: return one short, buyer-friendly English SKU name.\n'
                'Rules:\n'
                '1. Output only the SKU name with no explanation.\n'
                '2. Keep it within 30 characters.\n'
                '3. Keep only key variant, model, or size details.\n'
                '4. Do not use banned marketing terms.\n\n'
                f'Product title: {title}\n'
                f'Product description: {description[:600]}\n'
                f'Original SKU name: {raw_name}\n'
                f'Dimension hint: {dimension_hint or "none"}\n'
                f'Site: PH\n'
                f'Result length must be <= {max_length} characters.'
            )
            self._set_last_prompt_trace(mode='sku', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables=variables, bundle=bundle)
        else:
            prompt = (
                '你是 Shopee 台湾站的 SKU 规格命名助手。\n'
                '目标：输出一个清晰易懂、繁体中文、适合买家阅读的 SKU 名称。\n'
                '规则：\n'
                '1. 只输出一个 SKU 名称，不要解释。\n'
                '2. 长度不能超过 30 个字符。\n'
                '3. 不要重复整条商品标题，只保留必要规格、型号、尺寸信息。\n'
                '4. 不要使用营销词或禁用词。\n'
                '5. 若原规格已经足够清晰，可只做繁体化和精简。\n\n'
                f'商品标题：{title}\n'
                f'商品描述：{description[:600]}\n'
                f'原始 SKU 名称：{raw_name}\n'
                f'尺寸提示：{dimension_hint or "无"}\n'
                f'站点：TW\n'
                f'返回结果必须 <= {max_length} 字。'
            )
            self._set_last_prompt_trace(mode='sku', prompt=prompt, prompt_source='hardcoded_fallback', template_version='legacy', rendered_variables=variables, bundle=bundle)
        result = self._call_llm_text(prompt, max_tokens=120, timeout=90)
        candidate = self._sanitize_sku_listing_name(result or fallback_name, max_length=max_length, forbidden_terms=forbidden_terms)
        is_valid, _ = self._validate_sku_listing_name(candidate, existing_names, max_length=max_length, forbidden_terms=forbidden_terms)
        if not is_valid:
            candidate = self._sanitize_sku_listing_name(fallback_name, max_length=max_length, forbidden_terms=forbidden_terms)
        candidate = self._ensure_unique_sku_name(candidate, existing_names, max_length=max_length)
        return candidate

    def build_sku_listing_names(self, product: ProductData, max_length: int = 30, bundle: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        bundle = bundle or self._resolve_runtime_bundle(product)
        raw_skus = product.get('db_skus') if isinstance(product.get('db_skus'), list) else product.get('skus')
        skus: List[Dict[str, Any]] = []
        if isinstance(raw_skus, list):
            typed_skus = cast(List[Any], raw_skus)
            skus = [cast(Dict[str, Any], sku) for sku in typed_skus if isinstance(sku, dict)]

        existing_names: List[str] = []
        optimized: List[Dict[str, Any]] = []
        for sku in skus:
            optimized_name = self._generate_sku_listing_name(product, sku, existing_names, max_length=max_length, bundle=bundle)
            existing_names.append(optimized_name)
            optimized.append({
                'sku_id': sku.get('id'),
                'sku_name': sku.get('sku_name') or sku.get('name'),
                'shopee_sku_name': optimized_name,
            })
        return optimized

    def _generate_title(self, product: ProductData, info: Dict[str, Any], bundle: Dict[str, Any]) -> str:
        forbidden_terms = self._active_forbidden_terms(bundle)
        prompt = self._build_title_prompt(product, info, bundle)
        result = self._call_llm_text(prompt, max_tokens=300, timeout=120)
        if not result:
            return normalize_text(product.get('optimized_title') or product.get('title'))
        return self._sanitize_title(result, forbidden_terms=forbidden_terms)

    def _generate_description(self, product: ProductData, info: Dict[str, Any], bundle: Dict[str, Any]) -> Optional[str]:
        forbidden_terms = self._active_forbidden_terms(bundle)
        prompt = self._build_desc_prompt(product, info, bundle)
        result = self._call_llm_text(prompt, max_tokens=4500, timeout=180)
        if not result:
            return None
        title = normalize_text(product.get('optimized_title') or product.get('title'))
        # The model is asked to front-load keywords, but we still enforce a
        # deterministic repair pass before final validation for batch stability.
        return self._reinforce_description_head(result, title, forbidden_terms=forbidden_terms, listing_language=self._listing_language(bundle))

    def generate_optimized_content(self, product: ProductData) -> Optional[ProductData]:
        bundle = self._resolve_runtime_bundle(product)
        forbidden_terms = self._active_forbidden_terms(bundle)
        info = self._extract_product_info(product)

        optimized_title = self._generate_title(product, info, bundle)
        title_ok, title_error = self._validate_title(optimized_title, forbidden_terms=forbidden_terms, bundle=bundle)
        if not title_ok:
            fallback_title = normalize_text(product.get('optimized_title') or product.get('title'))
            optimized_title = self._sanitize_title(fallback_title, forbidden_terms=forbidden_terms)
            title_ok, title_error = self._validate_title(optimized_title, forbidden_terms=forbidden_terms, bundle=bundle)
            if not title_ok:
                print(f'标题校验失败: {title_error}')
                return None

        prompt_product = dict(product)
        prompt_product['optimized_title'] = optimized_title
        prompt_product.update(bundle.get('site_context') or {})
        optimized_description = self._generate_description(prompt_product, info, bundle)
        if not optimized_description:
            return None

        desc_ok, desc_error = self._validate_description(optimized_description, optimized_title, forbidden_terms=forbidden_terms, bundle=bundle)
        if not desc_ok:
            print(f'描述校验失败: {desc_error}')
            return None

        optimized_skus = self.build_sku_listing_names({**product, **(bundle.get('site_context') or {}), 'optimized_title': optimized_title, 'optimized_description': optimized_description}, bundle=bundle)

        return {
            'optimized_title': optimized_title,
            'optimized_description': optimized_description,
            'optimized_skus': optimized_skus,
        }

    def get_products_to_optimize(self, limit: int = 10, force: bool = False) -> List[ProductData]:
        cur = self.conn.cursor()

        if force:
            cur.execute(
                """
                SELECT product_id_new, alibaba_product_id, title, skus,
                       optimized_title, optimized_description
                FROM products
                WHERE (optimized_description IS NOT NULL AND optimized_description != '')
                  AND LENGTH(optimized_description) >= 100
                ORDER BY updated_at ASC
                LIMIT %s
                """,
                (limit,),
            )
        else:
            checker = ConsistencyChecker()
            issues = checker.check_all_products()
            checker.close()

            item_nos = [item['product']['item_no'] for item in issues[:limit]]
            if not item_nos:
                return []

            placeholders = ','.join(['%s'] * len(item_nos))
            cur.execute(
                f"""
                SELECT product_id_new, alibaba_product_id, title, skus,
                       optimized_title, optimized_description
                FROM products
                WHERE product_id_new IN ({placeholders})
                """,
                tuple(item_nos),
            )

        products: List[ProductData] = []
        for row in cur.fetchall():
            products.append({
                'item_no': row[0],
                'alibaba_id': row[1],
                'title': row[2],
                'skus': row[3],
                'optimized_title': row[4],
                'optimized_description': row[5],
            })

        cur.close()
        return products

    def optimize_one(self, item_no: str) -> Dict[str, Any]:
        product = self._fetch_product_by_item_no(item_no)
        if not product:
            return {'status': 'error', 'message': f'商品 {item_no} 未找到'}

        new_content = self.generate_optimized_content(product)
        if not new_content:
            return {'status': 'error', 'message': 'LLM调用失败或生成结果未通过校验'}

        self._update_product_content(item_no, new_content)
        return {
            'status': 'success',
            'item_no': item_no,
            'title_length': len(new_content['optimized_title']),
            'desc_length': len(new_content['optimized_description']),
        }

    def optimize_batch(self, limit: int = 10, force: bool = False) -> Dict[str, Any]:
        products = self.get_products_to_optimize(limit=limit, force=force)
        if not products:
            return {'status': 'success', 'message': '没有需要优化的商品', 'count': 0}

        success_count = 0
        failed: List[Dict[str, Any]] = []
        for product in products:
            new_content = self.generate_optimized_content(product)
            if not new_content:
                failed.append({'item_no': product['item_no'], 'error': 'LLM调用失败或生成结果未通过校验'})
                continue
            self._update_product_content(product['item_no'], new_content)
            success_count += 1

        return {
            'status': 'success',
            'total': len(products),
            'success': success_count,
            'failed': failed,
        }

    def _resolve_excel_columns(self, sheet: Worksheet) -> Dict[str, str]:
        row1: Dict[str, str] = {
            str(cell.column_letter): normalize_text(cell.value)
            for cell in cast(List[Any], sheet[1])
        }
        row3: Dict[str, str] = {
            str(cell.column_letter): normalize_text(cell.value)
            for cell in cast(List[Any], sheet[3])
        }

        resolved: Dict[str, str] = {}
        for key, candidates in EXCEL_COLUMN_KEYS.items():
            for column, value in row3.items():
                if value in candidates:
                    resolved[key] = column
                    break
            if key in resolved:
                continue
            for column, value in row1.items():
                if value in candidates:
                    resolved[key] = column
                    break

        missing = [key for key in EXCEL_COLUMN_KEYS if key not in resolved]
        if missing:
            raise ValueError(f'Excel 模板缺少必要列: {missing}')
        return resolved

    def _build_excel_product(self, sheet: Worksheet, row_index: int, columns: Dict[str, str]) -> ProductData:
        global_id = normalize_text(sheet[f"{columns['global_id']}{row_index}"].value)
        product_name = normalize_text(sheet[f"{columns['product_name']}{row_index}"].value)
        product_description = normalize_text(sheet[f"{columns['product_description']}{row_index}"].value)
        parent_sku = normalize_text(sheet[f"{columns['parent_sku']}{row_index}"].value)

        db_product = self._fetch_product_for_excel_row(parent_sku, global_id) or {}
        return {
            'item_no': parent_sku or normalize_text(db_product.get('item_no')),
            'alibaba_id': global_id or normalize_text(db_product.get('alibaba_id')),
            'title': product_name or normalize_text(db_product.get('title')),
            'optimized_title': product_name or normalize_text(db_product.get('optimized_title')),
            'optimized_description': product_description or normalize_text(db_product.get('optimized_description')),
            'skus': db_product.get('skus') or [],
        }

    def optimize_excel_template(
        self,
        input_path: str,
        output_path: Optional[str] = None,
        sheet_name: Optional[str] = None,
        start_row: int = 7,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        source = Path(input_path)
        if not source.exists():
            return {'status': 'error', 'message': f'文件不存在: {input_path}'}

        target = Path(output_path) if output_path else source.with_name(f'{source.stem}.optimized{source.suffix}')
        with self._open_workbook(source) as workbook:
            sheet = cast(Optional[Worksheet], workbook[sheet_name] if sheet_name else workbook.active)
            if sheet is None:
                return {'status': 'error', 'message': '未找到可用工作表'}
            columns = self._resolve_excel_columns(sheet)

            processed = 0
            success_count = 0
            failures: List[Dict[str, Any]] = []

            for row_index in range(start_row, sheet.max_row + 1):
                global_id = normalize_text(sheet[f"{columns['global_id']}{row_index}"].value)
                product_name = normalize_text(sheet[f"{columns['product_name']}{row_index}"].value)
                product_description = normalize_text(sheet[f"{columns['product_description']}{row_index}"].value)
                parent_sku = normalize_text(sheet[f"{columns['parent_sku']}{row_index}"].value)

                if not any([global_id, product_name, product_description, parent_sku]):
                    continue
                if limit is not None and processed >= limit:
                    break

                processed += 1

                try:
                    product = self._build_excel_product(sheet, row_index, columns)
                    if not normalize_text(product.get('title')):
                        raise ValueError('缺少商品名称，无法生成优化内容')

                    content = self.generate_optimized_content(product)
                    if not content:
                        raise ValueError('LLM调用失败或生成结果未通过校验')

                    sheet[f"{columns['product_name']}{row_index}"] = content['optimized_title']
                    sheet[f"{columns['product_description']}{row_index}"] = content['optimized_description']
                    sheet[f"{columns['failure_reason']}{row_index}"] = ''
                    success_count += 1
                except Exception as exc:
                    error = str(exc)
                    sheet[f"{columns['failure_reason']}{row_index}"] = error[:255]
                    failures.append({
                        'row': row_index,
                        'global_id': global_id,
                        'parent_sku': parent_sku,
                        'error': error,
                    })

            workbook.save(target)

        return {
            'status': 'success',
            'input': str(source),
            'output': str(target),
            'sheet': sheet.title,
            'processed': processed,
            'success': success_count,
            'failed': failures,
        }


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='Listing 批量优化工具')
    parser.add_argument('--check', action='store_true', help='只检查一致性')
    parser.add_argument('--optimize', action='store_true', help='执行数据库商品优化')
    parser.add_argument('--force', action='store_true', help='强制重新优化数据库商品')
    parser.add_argument('--limit', type=int, default=10, help='批量处理数量')
    parser.add_argument('--item-no', type=str, help='指定商品货号')
    parser.add_argument('--input-xlsx', type=str, help='Shopee 批量更新 Excel 模板路径')
    parser.add_argument('--output-xlsx', type=str, help='优化后 Excel 输出路径')
    parser.add_argument('--sheet-name', type=str, help='指定工作表名称')
    parser.add_argument('--start-row', type=int, default=7, help='Excel 数据起始行')

    args = parser.parse_args()

    if args.check:
        print('=== 一致性检查 ===')
        checker = ConsistencyChecker()
        issues = checker.check_all_products()
        checker.close()
        print(f'发现 {len(issues)} 个问题商品\n')
        for item in issues[:10]:
            print(f"货号: {item['product']['item_no']}")
            print(f"问题: {item['issues']}")
            print()

    elif args.input_xlsx:
        print(f'=== 优化 Excel 模板: {args.input_xlsx} ===')
        optimizer = ListingBatchOptimizer()
        result = optimizer.optimize_excel_template(
            input_path=args.input_xlsx,
            output_path=args.output_xlsx,
            sheet_name=args.sheet_name,
            start_row=args.start_row,
            limit=args.limit,
        )
        optimizer.close()
        print(result)

    elif args.item_no:
        print(f'=== 优化商品 {args.item_no} ===')
        optimizer = ListingBatchOptimizer()
        result = optimizer.optimize_one(args.item_no)
        optimizer.close()
        print(result)

    elif args.optimize:
        print(f'=== 批量优化数据库商品 (limit={args.limit}, force={args.force}) ===')
        optimizer = ListingBatchOptimizer()
        result = optimizer.optimize_batch(limit=args.limit, force=args.force)
        optimizer.close()
        print(result)

    else:
        parser.print_help()
